import pandas as pd
import openpyxl
import zipfile
import os

def compress_folder(folder_path, zip_file_name):
  """
  Compresses a folder into a ZIP archive.

  Args:
    folder_path: The path to the folder to compress.
    zip_file_name: The name of the ZIP file to create.
  """

  with zipfile.ZipFile(zip_file_name, 'w', zipfile.ZIP_DEFLATED) as zip_file:
    for root, dirs, files in os.walk(folder_path):
      for file in files:
        zip_file.write(os.path.join(root, file), os.path.relpath(os.path.join(root, file), folder_path))


def main():

    # cents.xlsx is pre-downloaded
    df = pd.read_excel('cents.xlsx')

    # Filtering for columns
    columns_of_interest = [
        'Number 7000',
        'Vote #2 Coloration: Verdigris',
        'Vote #2\nColoration: Red',
        'Vote #2\nColoration: Gold',
        'Vote #2\nColoration: Desert',
        'Vote #2\nColoration: Obsidian',
        'Vote #2\nLustrous',
        'Vote #2\nStriated',
        'Vote #2\nAbraded',
        'Vote #2\nContaminated',
        'Vote #2\nVandalized',
        'Vote #2\nFingerprint',
        'Vote #2\nGhost',
        'Vote #2\nAssassin',
        'Vote #2\nFaceless',
        'Inscription ID',
        'Image URL'
    ]
    filtered = df[columns_of_interest]

    # Rename columns for readability
    filtered.rename(columns={
        'Number 7000': 'Number',
        'Vote #2 Coloration: Verdigris': 'Verdigris',
        'Vote #2\nColoration: Red': 'Red',
        'Vote #2\nColoration: Gold': 'Gold',
        'Vote #2\nColoration: Desert': 'Desert',
        'Vote #2\nColoration: Obsidian': 'Obsidian',
        'Vote #2\nLustrous': 'Lustrous',
        'Vote #2\nStriated': 'Striated',
        'Vote #2\nAbraded': 'Abraded',
        'Vote #2\nContaminated': 'Contaminated',
        'Vote #2\nVandalized': 'Vandalized',
        'Vote #2\nFingerprint': 'Fingerprint',
        'Vote #2\nGhost': 'Ghost',
        'Vote #2\nAssassin': 'Assassin',
        'Vote #2\nFaceless': 'Faceless',
    }, inplace=True)

    # Prepare for sheets generation
    sheets_names = [
        'Verdigris',
        'Red',
        'Gold',
        'Desert',
        'Obsidian'
    ]

    sheets = {}

    # Loop generation of all notables
    for sheet_name in sheets_names:
        df_tmp = filtered[[
            'Number',
            sheet_name,
            'Inscription ID',
            'Image URL'
        ]]

        # Filter for notables and fulls
        notables = df_tmp[df_tmp[sheet_name] == 'Notable']
        fulls = df_tmp[df_tmp[sheet_name] == 'Full']
        # Rename column name
        notables.rename(columns={'Image URL': 'IMAGE'}, inplace=True)
        fulls.rename(columns={'Image URL': 'IMAGE'}, inplace=True)

        # Convert to dictionary
        notable_records = notables.to_dict('records')
        full_records = fulls.to_dict('records')

        
        # Loop through records to create xlsx file

        # Create a new workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # Iterate through the data, inserting into the worksheet

        # Notables
        row = 1
        col = 1
        for item in notable_records:
            worksheet.cell(row=row, column=col).value = item['Number']
            worksheet.cell(row=row+1, column=col).value = '=IMAGE("{0}")'.format(item['IMAGE'])
            worksheet.cell(row=row+2, column=col).value = item['Inscription ID']
            worksheet.cell(row=row+3, column=col).value = 1

            col += 1
            if col > 6:
                row += 5
                col = 1
        
        # Adjust size for viewability
        for scol in 'ABCDEF':
            worksheet.column_dimensions[scol].width = 35
        for srow in range(1, row+1):
            if srow % 5 == 2:
                worksheet.row_dimensions[srow].height = 200
            if srow % 5 == 0:
                worksheet.row_dimensions[srow].height = 100

        # Save the workbook
        workbook.save(f'notables/{sheet_name}_notables.xlsx')


        # Create a new workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Fulls
        row = 1
        col = 1
        for item in full_records:
            worksheet.cell(row=row, column=col).value = item['Number']
            worksheet.cell(row=row+1, column=col).value = '=IMAGE("{0}")'.format(item['IMAGE'])
            worksheet.cell(row=row+2, column=col).value = item['Inscription ID']
            worksheet.cell(row=row+3, column=col).value = 1

            col += 1
            if col > 6:
                row += 5
                col = 1
        
        # Adjust size for viewability
        for scol in 'ABCDEF':
            worksheet.column_dimensions[scol].width = 35
        for srow in range(1, row+1):
            if srow % 5 == 2:
                worksheet.row_dimensions[srow].height = 200
            if srow % 5 == 0:
                worksheet.row_dimensions[srow].height = 100

        # Save the workbook
        workbook.save(f'fulls/{sheet_name}_fulls.xlsx')

    # Compress both notables and fulls into zip archive
    compress_folder("notables", "notables.zip")
    compress_folder("fulls", "fulls.zip")


    other_sheet_names = [
        'Lustrous',
        'Striated',
        'Abraded',
        'Contaminated',
        'Vandalized',
        'Fingerprint',
        'Ghost',
        'Assassin',
        'Faceless'
    ]

    other_sheets = {}

    for sheet_name in other_sheet_names:
        df_tmp = filtered[[
            'Number',
            sheet_name,
            'Inscription ID',
            'Image URL'
        ]]

        # Filter for yes
        yes = df_tmp[df_tmp[sheet_name] == 'Yes']

        # Rename column name
        yes.rename(columns={'Image URL': 'IMAGE'}, inplace=True)

        # Convert to dictionary
        yes_records = yes.to_dict('records')

        # Loop through records to create xlsx file

        # Create a new workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # Iterate through the data, inserting into the worksheet

        # Yeses
        row = 1
        col = 1
        for item in yes_records:
            worksheet.cell(row=row, column=col).value = item['Number']
            worksheet.cell(row=row+1, column=col).value = '=IMAGE("{0}")'.format(item['IMAGE'])
            worksheet.cell(row=row+2, column=col).value = item['Inscription ID']
            worksheet.cell(row=row+3, column=col).value = 1

            col += 1
            if col > 6:
                row += 5
                col = 1
        
        # Adjust size for viewability
        for scol in 'ABCDEF':
            worksheet.column_dimensions[scol].width = 35
        for srow in range(1, row+1):
            if srow % 5 == 2:
                worksheet.row_dimensions[srow].height = 200
            if srow % 5 == 0:
                worksheet.row_dimensions[srow].height = 100

        # Save the workbook
        workbook.save(f'others/{sheet_name}.xlsx')

    # Compress both others folder into zip archive
    compress_folder("others", "others.zip")

if __name__ == "__main__":
    main()